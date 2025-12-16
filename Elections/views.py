from django.db import models
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
import csv
from .forms import ElectionForm, PositionForm, CandidateForm, VoterForm
from django.db.models import Count
from io import TextIOWrapper
from django.db import IntegrityError, transaction
from django.utils.timezone import make_aware
from django.db.models import Count
import logging
from openpyxl import load_workbook
from datetime import datetime
from .models import Election, Candidate, Vote, Position, Voter

# Create your views here.

@login_required
def dashboard(request):
    elections = Election.objects.filter(created_by=request.user) | Election.objects.filter(is_active=True)
    elections = elections.distinct().order_by('-created_at')
    total_voters = Voter.objects.count()
    active_count = Election.objects.filter(is_active=True).count()
    return render(request, 'dashboard.html', {
        'elections': elections,
        'total_voters': total_voters,
        'active_count': active_count,
    })

@login_required
def create_election(request):
    if request.method == "POST":
        title = request.POST.get("title")
        name = request.POST.get("name")
        description = request.POST.get("description")

        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        start_time = request.POST.get("start_time")
        end_time = request.POST.get("end_time")

        start_time = timezone.make_aware(
            datetime.fromisoformat(start_time),
            timezone.get_current_timezone()
        )

        end_time = timezone.make_aware(
            datetime.fromisoformat(end_time),
            timezone.get_current_timezone()
        )
        election = Election.objects.create(
            title=title,
            name=name,
            description=description,
            start_date=start_date,
            end_date=end_date,
            start_time=start_time,
            end_time=end_time,
            created_by=request.user
        )

        messages.success(request, "Election created successfully.")
        return redirect("Election:dashboard")

    return render(request, "create_election.html")
@login_required
def create_position(request):
    if request.method == 'POST':
        form = PositionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Position created.')
            return redirect('elections:manage_candidates')
    else:
        form = PositionForm()
    return render(request, 'create_position.html', {'form': form})



# def list_elections(request):
#     elections = Election.objects.order_by("-start_time")
#     return render(request, "list.html", {"elections": elections})

@login_required
def manage_candidates(request):
    positions = Position.objects.select_related('election').all()
    return render(request, 'manage_candidates.html', {'positions': positions})

@login_required
def add_candidate(request):
    if request.method == 'POST':
        form = CandidateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Candidate added.')
            return redirect('elections:manage_candidates')
    else:
        form = CandidateForm()
    return render(request, 'add_candidate.html', {'form': form})

@login_required
def add_voter(request):
    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        email = request.POST.get('email')
        election_id = request.POST.get('election')

        if not (full_name and email and election_id):
            messages.error(request, "All fields are required.")
        else:
            election = Election.objects.get(id=election_id)
            voter, created = Voter.objects.get_or_create(
                election=election,
                full_name=full_name,
                email=email,
                added_by=request.user
            )
            messages.success(request, "Voter added successfully.")
            return redirect('Election:dashboard')

    elections = Election.objects.all()
    return render(request, 'add_voters.html', {
        'elections': elections
    })

@login_required
def import_voters(request):
    elections = Election.objects.all()

    # Values we will send back to template
    voters = None
    import_summary = None

    if request.method == 'POST':
        uploaded_file = request.FILES.get('csv_file')
        election_id = request.POST.get('election')

        if not uploaded_file or not election_id:
            messages.error(request, "Election and file are required.")
            return redirect('elections:import_voters')

        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            messages.error(request, "Invalid election selected.")
            return redirect('elections:import_voters')

        imported = 0
        skipped = 0
        filename = uploaded_file.name.lower()

        # =========================
        # ✅ CSV IMPORT
        # =========================
        if filename.endswith('.csv'):
            data = TextIOWrapper(uploaded_file.file, encoding='utf-8', errors='ignore')
            reader = csv.DictReader(data)

            for row in reader:
                email = (row.get('email') or '').strip()
                full_name = (row.get('full_name') or row.get('name') or '').strip()

                if not email:
                    skipped += 1
                    continue

                voter, created = Voter.objects.get_or_create(
                    election=election,
                    email=email,
                    defaults={
                        'full_name': full_name,
                        'added_by': request.user
                    }
                )

                if created:
                    imported += 1
                else:
                    skipped += 1

        # =========================
        # ✅ EXCEL (.xlsx) IMPORT
        # =========================
        elif filename.endswith('.xlsx'):
            wb = load_workbook(uploaded_file)
            sheet = wb.active

            headers = [str(cell.value).lower().strip() for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))

                email = row_data.get('email')
                full_name = row_data.get('full_name') or row_data.get('name')

                if not email:
                    skipped += 1
                    continue

                voter, created = Voter.objects.get_or_create(
                    election=election,
                    email=str(email).strip(),
                    defaults={
                        'full_name': str(full_name).strip() if full_name else '',
                        'added_by': request.user
                    }
                )

                if created:
                    imported += 1
                else:
                    skipped += 1

        else:
            messages.error(request, "Only CSV or Excel (.xlsx) files are allowed.")
            return redirect('elections:import_voters')

        # ✅ fetch voters AFTER import
        voters = Voter.objects.filter(election=election).order_by('-id')

        import_summary = {
            'imported': imported,
            'skipped': skipped
        }

    return render(request, 'import_voters.html', {
        'elections': elections,
        'voters': voters,
        'import_summary': import_summary
    })
# @login_required
# def import_voters(request):
#     if request.method == 'POST' and request.FILES.get('csv_file'):
#         csv_file = request.FILES['csv_file']
#         try:
#             data = TextIOWrapper(csv_file.file, encoding='utf-8')
#         except Exception:
#             data = TextIOWrapper(csv_file, encoding='utf-8')
#         reader = csv.DictReader(data)
#         created = 0
#         for row in reader:
#             email = row.get('email') or row.get('Email') or ''
#             full_name = row.get('full_name') or row.get('name') or row.get('FullName') or ''
#             election_id = request.POST.get('election')
#             if not (email and election_id):
#                 continue
#             election = Election.objects.get(pk=int(election_id))
#             # avoid duplicates
#             obj, created_flag = Voter.objects.get_or_create(election=election, email=email, defaults={'full_name': full_name, 'added_by': request.user})
#             if created_flag:
#                 created += 1
#         messages.success(request, f'Imported {created} voters.')
#         return redirect('elections:dashboard')
#
#     elections = Election.objects.all()
#     return render(request, 'import_voters.html', {'elections': elections})

@login_required
def election_detail(request, pk):
    election = get_object_or_404(Election, pk=pk)
    positions = election.positions.prefetch_related('candidates').all()
    return render(request, 'election_detail.html', {'election': election, 'positions': positions})

@login_required
def vote_page(request, election_id):
    election = get_object_or_404(Election, pk=election_id)

    if not election.is_open():
        messages.error(request, 'This election is not open for voting.')
        return redirect('elections:dashboard')

    # ✅ get the Voter instance (NOT the User)
    try:
        voter = Voter.objects.get(election=election, email=request.user.email)
    except Voter.DoesNotExist:
        messages.error(request, 'You are not registered to vote in this election.')
        return redirect('elections:dashboard')

    positions = election.positions.prefetch_related('candidates').all()

    if request.method == 'POST':
        for pos in positions:
            key = f'choice_{pos.id}'
            cand_id = request.POST.get(key)

            if cand_id:
                candidate = get_object_or_404(
                    Candidate,
                    pk=int(cand_id),
                    position=pos
                )

                # ✅ FIX IS HERE
                Vote.objects.update_or_create(
                    election=election,
                    position=pos,
                    voter=voter,              # 👈 CHANGED
                    defaults={'candidate': candidate}
                )

        messages.success(request, 'Your vote has been recorded. Thank you!')
        return redirect('elections:dashboard')

    return render(
        request,
        'vote_page.html',
        {'election': election, 'positions': positions}
    )


@login_required
def results(request, election_id):
    election = get_object_or_404(Election, pk=election_id)

    # Get all positions for this election
    positions = election.positions.all()

    positions_with_votes = []

    for pos in positions:
        candidates_data = []

        # Loop through candidates in this position
        for c in pos.candidates.all():
            votes_count = Vote.objects.filter(
                election=election,
                position=pos,
                candidate=c
            ).count()

            candidates_data.append({
                'candidate': c,
                'votes_count': votes_count
            })

        positions_with_votes.append({
            'position': pos,
            'candidates': candidates_data
        })

    total_votes = Vote.objects.filter(election=election).count()
    registered_voters = Voter.objects.filter(election=election).count()
    turnout_percent = round(
        (total_votes / registered_voters * 100) if registered_voters else 0,
        2
    )

    return render(request, 'results.html', {
        'election': election,
        'positions_with_votes': positions_with_votes,
        'total_votes': total_votes,
        'registered_voters': registered_voters,
        'turnout_percent': turnout_percent,
    })



#
# @login_required
# def results(request, election_id):
#     election = get_object_or_404(Election, pk=election_id)
#     positions = election.positions.all().annotate(candidate_count=Count('candidates'))
#     # For each candidate, count votes
#     positions_with_votes = []
#     for pos in positions:
#         candidates = pos.candidates.annotate(votes_count=Count('candidate__vote', filter=models.Q(candidate__vote__position=pos)))
#         # simpler alternative: use Vote model aggregation
#         candidates_data= []
#         for c in pos.candidates.all():
#             votes_count = Vote.objects.filter(election=election, position=pos, candidate=c).count()
#             candidates_data.append({'candidate': c, 'votes_count': votes_count})
#         positions_with_votes.append({'position': pos, 'candidates': candidate=c})
#     total_votes = Vote.objects.filter(election=election).count()
#     registered_voters = Voter.objects.filter(election=election).count()
#     turnout_percent = round((total_votes / registered_voters * 100) if registered_voters else 0, 2)
#     return render(request, 'results.html', {
#         'election': election,
#         'positions_with_votes': positions_with_votes,
#         'total_votes': total_votes,
#         'registered_voters': registered_voters,
#         'turnout_percent': turnout_percent,
#     })



@login_required
def cast_vote(request, pk):
    election = get_object_or_404(Election, pk=pk)
    if not election.is_open():
        messages.error(request, "This election is not currently open.")
        return redirect("elections:detail", pk=pk)

    # check if user already voted
    if Vote.objects.filter(election=election, voter=request.user).exists():
        messages.error(request, "You have already voted in this election.")
        return redirect("elections:detail", pk=pk)

    if request.method == "POST":
        candidate_id = request.POST.get("candidate")
        candidate = get_object_or_404(Candidate, pk=candidate_id, election=election)
        Vote.objects.create(election=election, candidate=candidate, voter=request.user)
        messages.success(request, f"Vote cast for {candidate.name}. Thank you!")
        return redirect("elections:detail", pk=pk)

    return redirect("elections:detail", pk=pk)

@login_required
def delete_election(request, election_id):
    election = get_object_or_404(Election, id=election_id)

    if election.created_by != request.user:
        messages.error(request, "You are not allowed to delete this election.")
        return redirect("elections:dashboard")

    if request.method == "POST":
        election.delete()
        messages.success(request, "Election deleted successfully.")
        return redirect("elections:dashboard")

    return render(request, "confirm_delete.html", {"election": election})

@login_required
def delete_voter(request, voter_id):
    voter = get_object_or_404(Voter, id=voter_id)

    # ✅ CORRECT check
    if Vote.objects.filter(voter=voter).exists():
        messages.error(request, "You cannot delete a voter who has already voted.")
        return redirect("elections:import_voters")

    voter.delete()
    messages.success(request, "Voter deleted successfully.")
    return redirect("elections:import_voters")

@login_required
def delete_all_voters(request, election_id):
    election = get_object_or_404(Election, id=election_id)

    deleted_count, _ = Voter.objects.filter(
        election=election,
        vote__isnull=True   # 🔥 NOT voted
    ).delete()

    messages.success(
        request,
        f"Deleted {deleted_count} voters who had not voted."
    )
    return redirect("elections:import_voters")
